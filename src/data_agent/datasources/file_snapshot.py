"""Safe CSV/XLSX ingestion into immutable DuckDB datasource snapshots."""

from __future__ import annotations

import csv
import hashlib
import os
import re
import stat
import tempfile
import unicodedata
import zipfile
from collections.abc import Iterable, Sequence
from datetime import date, datetime
from decimal import Decimal
from enum import StrEnum
from pathlib import Path

import duckdb
from openpyxl import load_workbook

from data_agent.tools.schemas import (
    CatalogColumn,
    CatalogRelation,
    CatalogSnapshot,
)

from .models import DataSourceModel, NonBlankText


class FileSnapshotErrorCode(StrEnum):
    FILE_NOT_FOUND = "FILE_NOT_FOUND"
    UNSUPPORTED_FORMAT = "UNSUPPORTED_FORMAT"
    SIZE_LIMIT_EXCEEDED = "SIZE_LIMIT_EXCEEDED"
    UNSAFE_ARCHIVE = "UNSAFE_ARCHIVE"
    EMPTY_DATASET = "EMPTY_DATASET"
    SNAPSHOT_EXISTS = "SNAPSHOT_EXISTS"
    IMPORT_FAILED = "IMPORT_FAILED"


class FileSnapshotError(ValueError):
    def __init__(
        self,
        code: FileSnapshotErrorCode,
        message: str,
    ) -> None:
        self.code = code
        super().__init__(message)


class ImportedRelation(DataSourceModel):
    origin_file: NonBlankText
    origin_table: NonBlankText
    relation: NonBlankText
    row_count: int


class FileSnapshotResult(DataSourceModel):
    database_path: Path
    fingerprint: NonBlankText
    catalog: CatalogSnapshot
    relations: tuple[ImportedRelation, ...]


def _quote_identifier(value: str) -> str:
    return '"' + value.replace('"', '""') + '"'


def _failure_reason(
    error: BaseException,
    *,
    internal_path: Path | None = None,
    display_name: str | None = None,
) -> str:
    reason = " ".join(str(error).strip().split()) or type(error).__name__
    if internal_path is not None and display_name is not None:
        reason = reason.replace(str(internal_path), display_name)
    return reason[:800]


def _safe_identifier(value: str, *, fallback: str) -> str:
    normalized = unicodedata.normalize("NFKD", value)
    ascii_value = normalized.encode("ascii", "ignore").decode("ascii")
    identifier = re.sub(r"[^a-zA-Z0-9_]+", "_", ascii_value).strip("_").lower()
    if not identifier:
        identifier = fallback
    if identifier[0].isdigit():
        identifier = "t_" + identifier
    return identifier[:63]


def _deduplicate_identifier(base: str, used: set[str]) -> str:
    candidate = base
    suffix = 2
    while candidate in used:
        marker = f"_{suffix}"
        candidate = base[: 63 - len(marker)] + marker
        suffix += 1
    used.add(candidate)
    return candidate


def _spreadsheet_value(value: object) -> object:
    if value is None:
        return ""
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    if isinstance(value, Decimal):
        return str(value)
    return value


class FileSnapshotImporter:
    """Convert trusted upload-staging files into a bounded DuckDB snapshot."""

    def __init__(
        self,
        *,
        staging_root: str | Path | None = None,
        max_files: int = 20,
        max_file_bytes: int = 64 * 1024 * 1024,
        max_archive_entries: int = 10_000,
        max_archive_uncompressed_bytes: int = 512 * 1024 * 1024,
        max_archive_ratio: int = 200,
        max_tables: int = 100,
        max_rows_per_table: int = 2_000_000,
        max_total_rows: int = 5_000_000,
        max_columns: int = 4_096,
    ) -> None:
        if max_rows_per_table < 1 or max_total_rows < max_rows_per_table:
            raise ValueError(
                "row limits must be positive and the total limit must be "
                "at least the per-table limit"
            )
        self._staging_root = (
            Path(staging_root).expanduser().resolve(strict=True)
            if staging_root is not None
            else None
        )
        self._max_files = max_files
        self._max_file_bytes = max_file_bytes
        self._max_archive_entries = max_archive_entries
        self._max_archive_uncompressed_bytes = max_archive_uncompressed_bytes
        self._max_archive_ratio = max_archive_ratio
        self._max_tables = max_tables
        self._max_rows_per_table = max_rows_per_table
        self._max_total_rows = max_total_rows
        self._max_columns = max_columns

    def _validate_source(self, source: str | Path) -> Path:
        display_name = Path(source).name or "upload"
        try:
            path = Path(source).expanduser().resolve(strict=True)
        except FileNotFoundError as exc:
            raise FileSnapshotError(
                FileSnapshotErrorCode.FILE_NOT_FOUND,
                f"file {display_name!r} was not found",
            ) from exc
        if not path.is_file():
            raise FileSnapshotError(
                FileSnapshotErrorCode.FILE_NOT_FOUND,
                f"file {display_name!r} is not a regular file",
            )
        if (
            self._staging_root is not None
            and not path.is_relative_to(self._staging_root)
        ):
            raise FileSnapshotError(
                FileSnapshotErrorCode.FILE_NOT_FOUND,
                f"file {display_name!r} is outside the authorized upload directory",
            )
        if path.stat().st_size > self._max_file_bytes:
            raise FileSnapshotError(
                FileSnapshotErrorCode.SIZE_LIMIT_EXCEEDED,
                (
                    f"file {display_name!r} is {path.stat().st_size:,} bytes; "
                    f"the per-file limit is {self._max_file_bytes:,} bytes"
                ),
            )
        if path.suffix.lower() not in {".csv", ".xlsx"}:
            raise FileSnapshotError(
                FileSnapshotErrorCode.UNSUPPORTED_FORMAT,
                (
                    f"file {display_name!r} has unsupported format "
                    f"{path.suffix or '(none)'}; only CSV and XLSX are supported"
                ),
            )
        return path

    def _validate_xlsx_archive(self, path: Path) -> None:
        if not zipfile.is_zipfile(path):
            raise FileSnapshotError(
                FileSnapshotErrorCode.UNSAFE_ARCHIVE,
                "XLSX upload is not a valid ZIP-based workbook",
            )
        with zipfile.ZipFile(path) as archive:
            entries = archive.infolist()
            if len(entries) > self._max_archive_entries:
                raise FileSnapshotError(
                    FileSnapshotErrorCode.UNSAFE_ARCHIVE,
                    "XLSX archive contains too many entries",
                )
            total_uncompressed = 0
            for entry in entries:
                total_uncompressed += entry.file_size
                if (
                    entry.file_size > 0
                    and (
                        entry.compress_size == 0
                        or entry.file_size
                        > entry.compress_size * self._max_archive_ratio
                    )
                ):
                    raise FileSnapshotError(
                        FileSnapshotErrorCode.UNSAFE_ARCHIVE,
                        "XLSX archive has an unsafe compression ratio",
                    )
                lowered = entry.filename.lower()
                if any(
                    marker in lowered
                    for marker in (
                        "vbaproject.bin",
                        "/activex/",
                        "/embeddings/",
                    )
                ):
                    raise FileSnapshotError(
                        FileSnapshotErrorCode.UNSAFE_ARCHIVE,
                        "XLSX archive contains active or embedded content",
                    )
            if total_uncompressed > self._max_archive_uncompressed_bytes:
                raise FileSnapshotError(
                    FileSnapshotErrorCode.UNSAFE_ARCHIVE,
                    "XLSX archive expands beyond the safety limit",
                )

    @staticmethod
    def _digest_sources(paths: Sequence[Path]) -> str:
        digest = hashlib.sha256(b"file-snapshot-v1\0")
        for path in paths:
            digest.update(path.name.encode("utf-8"))
            digest.update(b"\0")
            with path.open("rb") as stream:
                for chunk in iter(lambda: stream.read(1024 * 1024), b""):
                    digest.update(chunk)
            digest.update(b"\0")
        return "sha256:" + digest.hexdigest()

    def _xlsx_csv_files(
        self,
        path: Path,
        *,
        temporary_directory: Path,
    ) -> Iterable[tuple[str, Path]]:
        self._validate_xlsx_archive(path)
        workbook = load_workbook(
            path,
            read_only=True,
            data_only=True,
            keep_links=False,
        )
        try:
            for index, worksheet in enumerate(workbook.worksheets, start=1):
                if worksheet.max_column > self._max_columns:
                    raise FileSnapshotError(
                        FileSnapshotErrorCode.SIZE_LIMIT_EXCEEDED,
                        "worksheet exceeds the column limit",
                    )
                if worksheet.max_row > self._max_rows_per_table + 1:
                    raise FileSnapshotError(
                        FileSnapshotErrorCode.SIZE_LIMIT_EXCEEDED,
                        "worksheet exceeds the row limit",
                    )
                rows = worksheet.iter_rows(values_only=True)
                header = next(rows, None)
                if header is None or not any(value is not None for value in header):
                    continue
                if len(header) > self._max_columns:
                    raise FileSnapshotError(
                        FileSnapshotErrorCode.SIZE_LIMIT_EXCEEDED,
                        "worksheet exceeds the column limit",
                    )
                used_headers: set[str] = set()
                headers = tuple(
                    _deduplicate_identifier(
                        _safe_identifier(
                            str(value or ""),
                            fallback=f"column_{column_index}",
                        ),
                        used_headers,
                    )
                    for column_index, value in enumerate(header, start=1)
                )
                handle = tempfile.NamedTemporaryFile(
                    mode="w",
                    encoding="utf-8",
                    newline="",
                    suffix=".csv",
                    prefix=f"xlsx-{index}-",
                    dir=temporary_directory,
                    delete=False,
                )
                temporary_path = Path(handle.name)
                try:
                    writer = csv.writer(handle)
                    writer.writerow(headers)
                    row_count = 0
                    for row in rows:
                        row_count += 1
                        if row_count > self._max_rows_per_table:
                            raise FileSnapshotError(
                                FileSnapshotErrorCode.SIZE_LIMIT_EXCEEDED,
                                "worksheet exceeds the row limit",
                            )
                        values = tuple(_spreadsheet_value(value) for value in row)
                        writer.writerow(
                            values[: len(headers)]
                            + ("",) * max(0, len(headers) - len(values))
                        )
                finally:
                    handle.close()
                yield worksheet.title, temporary_path
        finally:
            workbook.close()

    def import_files(
        self,
        sources: Sequence[str | Path],
        *,
        output_directory: str | Path,
        source_id: str,
        version: int,
        schema: str = "public",
    ) -> FileSnapshotResult:
        if not sources or len(sources) > self._max_files:
            raise FileSnapshotError(
                FileSnapshotErrorCode.SIZE_LIMIT_EXCEEDED,
                "file datasource must contain a bounded number of uploads",
            )
        paths = tuple(self._validate_source(source) for source in sources)
        if len({path.name for path in paths}) != len(paths):
            duplicates = sorted(
                name
                for name in {path.name for path in paths}
                if sum(path.name == name for path in paths) > 1
            )
            raise FileSnapshotError(
                FileSnapshotErrorCode.IMPORT_FAILED,
                "uploaded filenames must be unique; duplicates: "
                + ", ".join(duplicates),
            )
        fingerprint = self._digest_sources(paths)
        output_root = Path(output_directory).expanduser().resolve()
        output_root.mkdir(parents=True, exist_ok=True)
        safe_source_id = _safe_identifier(source_id, fallback="dataset")
        database_path = (
            output_root
            / f"{safe_source_id}-v{version}-{fingerprint[7:23]}.duckdb"
        )
        if database_path.exists():
            raise FileSnapshotError(
                FileSnapshotErrorCode.SNAPSHOT_EXISTS,
                "immutable file datasource snapshot already exists",
            )

        connection: duckdb.DuckDBPyConnection | None = None
        temporary_paths: list[Path] = []
        imported: list[ImportedRelation] = []
        used_tables: set[str] = set()
        total_rows = 0
        completed = False
        try:
            connection = duckdb.connect(
                str(database_path),
                config={
                    "allow_community_extensions": "false",
                    "autoload_known_extensions": "false",
                    "autoinstall_known_extensions": "false",
                },
            )
            connection.execute(f"CREATE SCHEMA {_quote_identifier(schema)}")

            candidates: list[tuple[Path, str, Path]] = []
            for source in paths:
                if source.suffix.lower() == ".csv":
                    try:
                        source.read_bytes()[: 64 * 1024].decode("utf-8-sig")
                    except UnicodeDecodeError as exc:
                        raise FileSnapshotError(
                            FileSnapshotErrorCode.IMPORT_FAILED,
                            (
                                f"file {source.name!r} is not valid UTF-8 CSV: "
                                f"{exc.reason} at byte {exc.start}"
                            ),
                        ) from exc
                    candidates.append((source, source.stem, source))
                    continue
                try:
                    with tempfile.TemporaryDirectory(
                        prefix="data-agent-xlsx-",
                        dir=output_root,
                    ) as temporary:
                        generated = tuple(
                            self._xlsx_csv_files(
                                source,
                                temporary_directory=Path(temporary),
                            )
                        )
                        for sheet_name, csv_path in generated:
                            persistent = output_root / (
                                f".{database_path.stem}-{len(temporary_paths)}.csv"
                            )
                            os.replace(csv_path, persistent)
                            temporary_paths.append(persistent)
                            candidates.append(
                                (
                                    source,
                                    f"{source.stem}_{sheet_name}",
                                    persistent,
                                )
                            )
                except FileSnapshotError as exc:
                    raise FileSnapshotError(
                        exc.code,
                        f"file {source.name!r} could not be imported: {exc}",
                    ) from exc
                except (OSError, ValueError, zipfile.BadZipFile) as exc:
                    raise FileSnapshotError(
                        FileSnapshotErrorCode.IMPORT_FAILED,
                        (
                            f"file {source.name!r} could not be imported: "
                            f"{_failure_reason(exc, internal_path=source, display_name=source.name)}"
                        )
                    ) from exc

            if not candidates or len(candidates) > self._max_tables:
                raise FileSnapshotError(
                    FileSnapshotErrorCode.EMPTY_DATASET,
                    "file datasource contains no importable tables",
                )

            for origin, origin_table, csv_path in candidates:
                table = _deduplicate_identifier(
                    _safe_identifier(
                        origin_table,
                        fallback=f"table_{len(imported) + 1}",
                    ),
                    used_tables,
                )
                qualified = (
                    f"{_quote_identifier(schema)}.{_quote_identifier(table)}"
                )
                try:
                    connection.execute(
                        f"""
                        CREATE TABLE {qualified} AS
                        SELECT *
                        FROM read_csv_auto(
                            ?,
                            header = true,
                            sample_size = -1,
                            strict_mode = true
                        )
                        LIMIT {self._max_rows_per_table + 1}
                        """,
                        [str(csv_path)],
                    )
                    row_count = int(
                        connection.execute(
                            f"SELECT COUNT(*) FROM {qualified}"
                        ).fetchone()[0]
                    )
                except duckdb.Error as exc:
                    raise FileSnapshotError(
                        FileSnapshotErrorCode.IMPORT_FAILED,
                        (
                            f"file {origin.name!r} could not be imported as "
                            f"table {schema}.{table}: "
                            f"{_failure_reason(exc, internal_path=csv_path, display_name=origin.name)}"
                        ),
                    ) from exc
                if row_count > self._max_rows_per_table:
                    raise FileSnapshotError(
                        FileSnapshotErrorCode.SIZE_LIMIT_EXCEEDED,
                        (
                            f"file {origin.name!r}, table {schema}.{table}, "
                            f"has {row_count:,} rows; "
                            "the per-table limit is "
                            f"{self._max_rows_per_table:,}"
                        ),
                    )
                total_rows += row_count
                if total_rows > self._max_total_rows:
                    raise FileSnapshotError(
                        FileSnapshotErrorCode.SIZE_LIMIT_EXCEEDED,
                        (
                            f"after file {origin.name!r}, the datasource has "
                            f"more than {total_rows:,} rows; "
                            "the total row limit is "
                            f"{self._max_total_rows:,}"
                        ),
                    )
                imported.append(
                    ImportedRelation(
                        origin_file=origin.name,
                        origin_table=origin_table,
                        relation=f"{schema}.{table}",
                        row_count=row_count,
                    )
                )

            column_rows = connection.execute(
                """
                SELECT table_schema, table_name, column_name, data_type, is_nullable
                FROM information_schema.columns
                ORDER BY table_schema, table_name, ordinal_position
                """
            ).fetchall()
            grouped: dict[str, list[CatalogColumn]] = {}
            for (
                table_schema,
                table_name,
                column_name,
                data_type,
                is_nullable,
            ) in column_rows:
                relation = f"{table_schema}.{table_name}"
                grouped.setdefault(relation, []).append(
                    CatalogColumn(
                        name=str(column_name),
                        data_type=str(data_type),
                        nullable=str(is_nullable).upper() == "YES",
                    )
                )
            catalog = CatalogSnapshot(
                schema_fingerprint=fingerprint,
                relations=tuple(
                    CatalogRelation(
                        relation=item.relation,
                        columns=tuple(grouped[item.relation]),
                        estimated_rows=item.row_count,
                    )
                    for item in imported
                ),
            )
            connection.close()
            connection = None
            database_path.chmod(stat.S_IRUSR | stat.S_IRGRP)
            result = FileSnapshotResult(
                database_path=database_path,
                fingerprint=fingerprint,
                catalog=catalog,
                relations=tuple(imported),
            )
            completed = True
            return result
        except FileSnapshotError:
            raise
        except (UnicodeError, OSError, ValueError, zipfile.BadZipFile, duckdb.Error) as exc:
            raise FileSnapshotError(
                FileSnapshotErrorCode.IMPORT_FAILED,
                "file datasource could not be imported safely",
            ) from exc
        finally:
            if connection is not None:
                connection.close()
            for temporary_path in temporary_paths:
                temporary_path.unlink(missing_ok=True)
            if database_path.exists() and not completed:
                database_path.chmod(stat.S_IWUSR | stat.S_IRUSR)
                database_path.unlink(missing_ok=True)


__all__ = [
    "FileSnapshotError",
    "FileSnapshotErrorCode",
    "FileSnapshotImporter",
    "FileSnapshotResult",
    "ImportedRelation",
]
